# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from datetime import datetime
import sys
import math
from functools import lru_cache

result_list = []
white_discount = ""
yellow_discount = ""
white_tmp_list = []
yellow_tmp_list = []


def now():
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    return current_time

@lru_cache(maxsize=None)
def create_list():
    if application_xls_child.input_file_label_text_field.get() != "":
        try:
            print(application_xls_child.input_file_label_text_field.get())
            global result_list
            print("начало создания списков", now())
            application_xls_child.time_start["text"] = "Время запуска: {}".format(now())
            application_xls_child.app.update()
            wb = load_workbook("{}".format(application_xls_child.input_file_label_text_field.get()), data_only=True)
            application_xls_child.app.update()
            wb = wb["{}".format(application_xls_child.output_file_label_text_field.get())]
            for row in wb.iter_rows(min_row=2,):
                for cell in row:
                    if str(cell.fill.start_color.index) == "00000000" and cell.value != None:
                        white_tmp_list.append(cell.value)
                    if str(cell.fill.start_color.index) != "00000000" and cell.value != None:
                        yellow_tmp_list.append(cell.value)
            application_xls_child.app.update()
            application_xls_child.progress["value"] += 20
            print("Конец, создания списков",now())
            thread_lists()
        except:
            application_xls_child.time_start["text"] = ("Выберите лист")
            application_xls_child.time_stop["text"] = ("для сканирования")
    else:
        application_xls_child.time_start["text"] = "Выберите файл"
        application_xls_child.time_stop["text"] = ""

def thread_lists():
    parse_white_list()
    parse_yellow_list()
    thread_write_file()

@lru_cache(maxsize=None)
def thread_write_file():
    print("thread write")
    tmp = ""
    global result_list
    global yellow_tmp_list
    global white_tmp_list
    result_list += yellow_tmp_list + white_tmp_list
    index_yourlist = 0
    print("mi doshli do syuda")
    for i in result_list:
        index_yourlist += 1
        if True:
            tmp += str(i) + ";"
        if index_yourlist % 5 == 0:
            tmp += "\n"
    print("zakonchili tmp")
    application_xls_child.progress["value"] += 20
    result_list = []
    yellow_tmp_list = []
    white_tmp_list = []
    file_ = application_xls_child.default_temp_path.get()
    file_obj = open(file_, "w")
    file_obj.write(tmp)
    file_obj.close()
    tmp = ""
    application_xls_child.progress["value"] += 20
    print("done", now())
    application_xls_child.time_stop["text"] = "Финишное время: {}".format(now())





@lru_cache(maxsize=None)
def parse_white_list():
    print("начало сканирования white_list", now())
    white_discount = application_xls_child.default_discount_white.get()
    index_white = 4
    for item in white_tmp_list:
        try:
            if white_tmp_list[index_white] == 0:
                incoming_price = white_tmp_list[index_white -1] / 1.2
                price_sale = incoming_price - ((incoming_price * float(white_discount)) / 100)
                white_tmp_list[index_white] = math.ceil(price_sale*100) / 100
                white_tmp_list[index_white -1] = math.ceil(incoming_price*100) / 100
                index_white += 5

            # if isinstance(white_tmp_list[index_white], float) != 0 or isinstance(white_tmp_list[index_white], int) != 0:
            else:
                incoming_price = white_tmp_list[index_white] / 1.2
                price_sale = incoming_price - ((incoming_price * float(white_discount)) / 100)
                white_tmp_list[index_white] = math.ceil(price_sale*100) / 100
                white_tmp_list[index_white - 1] = math.ceil(incoming_price*100) / 100
                index_white += 5

        except IndexError:
            pass
    application_xls_child.progress["value"] += 20
    print("конец сканирования white_list", now())

    # print("С обработкой", self.white_tmp_list)

    # print("Без обработки", self.yellow_tmp_list)


@lru_cache(maxsize=None)
def parse_yellow_list():
    print("начало сканирования yellow_list", now())
    yellow_discount = application_xls_child.default_discount_yellow.get()
    index_yellow = 4

    for item in yellow_tmp_list:
        try:
            if yellow_tmp_list[index_yellow] == 0:
                incoming_price = yellow_tmp_list[index_yellow -1] / 1.2
                price_sale = incoming_price - ((incoming_price * float(yellow_discount)) / 100)
                print(price_sale)
                yellow_tmp_list[index_yellow] = math.ceil(price_sale*100) / 100
                yellow_tmp_list[index_yellow - 1] = math.ceil(incoming_price * 100) / 100
                index_yellow += 5

            else:
                incoming_price = yellow_tmp_list[index_yellow] / 1.2
                price_sale = incoming_price - ((incoming_price * float(yellow_discount)) / 100)
                yellow_tmp_list[index_yellow] = math.ceil(price_sale*100) / 100
                yellow_tmp_list[index_yellow - 1] = math.ceil(incoming_price*100) / 100
                index_yellow += 5

        except IndexError:
            pass
    # print("С обработкой", self.yellow_tmp_list)
    application_xls_child.progress["value"] += 20
    print("конец сканирования yellow_list", now())





class Application_Xls():
    def __init__(self):
        self.app = tk.Tk()
        global white_discount
        global yellow_discount
        self.app.title("Parser xlsx")
        self.app.resizable(False, False)
        self.button_memory = ttk.Button(self.app, text="Исходный файл(excel)",command= lambda: self.inputfile(), style="my.TButton",)
        self.button_memory.grid(row=1, column=3, ipady=0, ipadx=25)
        self.app.bind("<Key>", self.key_enter)
        self.button_percent = ttk.Button(self.app, text="Файл назначения(txt)", command=lambda: self.outputfile(), style="my.TButton")
        self.button_percent.grid(row=2, column=3, ipady=0, ipadx=27)
        self.button_percent = ttk.Button(self.app, text="Пуск", command=lambda: create_list(), style="my.TButton")
        self.button_percent.grid(row=6, column=3, ipady=0, ipadx=27)
        self.default_incoming_file = tk.StringVar()
        self.default_incoming_file.set("")
        self.default_temp_path = tk.StringVar()
        self.default_temp_path.set("C:/Temp/out.txt")
        self.default_latter = tk.StringVar()
        self.default_latter.set(r"Лист1")
        self.default_discount_yellow = tk.StringVar()
        self.default_discount_yellow.set(25)
        self.default_discount_white = tk.StringVar()
        self.default_discount_white.set(35)
        self.input_file_label = ttk.Label(self.app, text="Выберите исходный файл: ")
        self.input_file_label.grid(row=1, column=0)
        self.input_file_label_text_field = ttk.Entry(self.app, width=30, textvariable=self.default_incoming_file)
        self.input_file_label_text_field.grid(row=1, column=1)
        self.output_file_label = ttk.Label(self.app, text="Выберите файл назначения: ")
        self.output_file_label.grid(row=2, column=0)
        self.output_file_label_text_field = ttk.Entry(self.app, width=30, textvariable=self.default_temp_path)
        self.output_file_label_text_field.grid(row=2, column=1)

        self.output_file_label = ttk.Label(self.app, text="Скидка желтая: ")
        self.output_file_label.grid(row=3, column=0)
        self.output_file_label_text_field = ttk.Entry(self.app, width=30, textvariable=self.default_discount_yellow)
        self.output_file_label_text_field.grid(row=3, column=1)

        self.output_file_label = ttk.Label(self.app, text="Скидка: ")
        self.output_file_label.grid(row=4, column=0)
        self.output_file_label_text_field = ttk.Entry(self.app, width=30, textvariable=self.default_discount_white)
        self.output_file_label_text_field.grid(row=4, column=1)
        self.time_start = ttk.Label(self.app, text="")
        self.time_start.grid(row=4, column=2, columnspan=2)
        self.time_stop = ttk.Label(self.app, text="")
        self.time_stop.grid(row=5, column=2, columnspan=2)
        self.output_file_label.grid(row=4, column=0)
        self.output_file_label = ttk.Label(self.app, text="По умолчанию=Лист1: ")
        self.output_file_label.grid(row=5, column=0)
        self.output_file_label_text_field = ttk.Entry(self.app, width=30, textvariable=self.default_latter)
        self.output_file_label_text_field.grid(row=5, column=1)
        self.progress = Progressbar(self.app, orient=HORIZONTAL, length=350, mode='determinate')
        self.progress.grid(row=6, column=0, columnspan=2)
        self.progress["value"] = 0

        # self.white_tmp_list = []
        # self.yellow_tmp_list = []
        self.skidka_zheltaya = float(self.default_discount_yellow.get())
        self.skidka_belaya = float(self.default_discount_white.get())

    def now(self):
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        print(current_time)
        return current_time

    # 2 кнопки для выбора файла
    def inputfile(self):
        self.filename_input = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                  filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        self.default_incoming_file.set("{}".format(self.filename_input))

    def outputfile(self):
        self.filename_output = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                  filetypes=(("txt files", "*.txt"), ("all files", "*.*")))
        self.default_temp_path.set("{}".format(self.filename_output))


    def key_enter(self, event):
        if event.keysym == "Return":
            create_list()
        if event.keysym == "Escape":
            sys.exit()
        else:
            pass
        print(event.keysym)


application_xls_child = Application_Xls()
application_xls_child.app.mainloop()

